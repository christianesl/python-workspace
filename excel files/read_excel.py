import openpyxl

book = openpyxl.load_workbook('data.xlsx')
worksheet = book.worksheets[0]

main_list = []

for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
	sub_lst = []
	for cell in row:
		if cell.value: 		
			sub_lst.append(str(cell.value))
	main_list.append(sub_lst)		
		

index = 0
for lst in main_list:
	print(lst)

